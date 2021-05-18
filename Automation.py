import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def get_spreadsheet(filename, name_of_sheet):
    # as xl is an alias of openpyxl
    wb = xl.load_workbook(filename)
    # a var contain the 1st sheet
    sheet = wb[name_of_sheet]
    # It is case sensitive

    # two method to get data from the cell in spreadsheet
    cell = sheet['A1']
    cell2 = sheet['b1']
    cell3 = sheet['C1']
    # cell = sheet.cell(1,1)

    # .value to get the value in the spreadsheet
    print('*'*53)
    print('EXCEL INFOR'.center(53))
    print('*'*53)
    print(str(cell.value).ljust(20),str(cell2.value).ljust(20),str(cell3.value).rjust(10))
    # print(cell2.value)
    # print(cell3.value)
    print('*'*53)
    # the number of rows in spreadsheet
    for row in range(2, sheet.max_row + 1):
        data = sheet.cell(row, 3) # (row, column)
        discount = data.value * 0.9
        new_row = sheet.cell(row, 4)
        new_row.value = round(discount, 2)

    # Draw the barchart
    values = Reference(sheet, min_row=2, max_row= sheet.max_row, min_col=4, max_col=4) # Get the data from row to row, and from column to column
    chart = BarChart() # declare the type of chart
    chart.add_data(values) # fill the data in the chart
    sheet.add_chart(chart, 'e2') # where to add the drew chart
    # method to save new modified spreaddsheet
    wb.save('transaction.xlsx')


get_spreadsheet('transactions.xlsx','Sheet1')